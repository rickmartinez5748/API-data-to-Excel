import requests
from openpyxl import Workbook

url = "https://disease.sh/v3/covid-19/countries"
response = requests.get(url)
data = response.json()

#Delimiting the data of COVID to countries of South America
south_america_countries=[]

for c in data:
    if c['continent'] == "South America":
        south_america_countries.append(c)
        

# Create Workbook 
wb = Workbook()
ws = wb.active
ws.title = "COVID Data"

# headers
headers = ['Country','Continent','Population','Total Cases','Total Deaths','Recovered','Active Cases','Tests Conducted','Cases per Million','Deaths per Million']
ws.append(headers)

#inserting information from RESTpoint selected
for cty in south_america_countries:
    ws.append([
        (cty['country']),
        (cty['continent']),
        int(cty['population']),
        int(cty['cases']),
        int(cty['deaths']),
        int(cty['recovered']),
        int(cty['active']),
        int(cty['tests']),
        int(cty['casesPerOneMillion']),
        int(cty['deathsPerOneMillion'])
    ])

#--------------------------------------------------------------------------------------------
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

#------------------------------------------------------------------------------------------
#Bar Chart
country=[]
total_cases=[]

#Reading values to create necessary vectors for image creation
for row in ws.iter_rows(min_row=2, values_only=True):
    country.append(row[0])       
    total_cases.append(row[3])   

# Create bar chart
plt.figure(figsize=(12, 6))
plt.bar(country, total_cases, color='blue')
plt.title("Total COVID-19 Cases in South America")
plt.xlabel("Country")
plt.ylabel("Total Cases")
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("total_cases_chart.png")
plt.close()

# Create a new sheet for the bar chart
ws_chart = wb.create_sheet("Bar Chart")

# Insert image in new sheet
img = Image("total_cases_chart.png")
ws_chart.add_image(img, "B2")

#--------------------------------------------------------------------------------------------
#Line Chart
country = []
cases_per_million = []

for row in ws.iter_rows(min_row=2, values_only=True):
    country.append(row[0])             
    cases_per_million.append(row[8]) 

# Create line plot
plt.figure(figsize=(12, 6))
plt.plot(country, cases_per_million, marker='o', color='green')
plt.title("COVID-19 Cases per Million - South America")
plt.xlabel("Country")
plt.ylabel("Cases per Million")
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig("cases_per_million_chart.png")
plt.close()

# Create a new sheet for the chart
ws_chart = wb.create_sheet("Line Chart")

# Insert image in new sheet
img = Image("cases_per_million_chart.png")
ws_chart.add_image(img, "B2")

#----------------------------------------------------------------------------------------------
#Pie Chart
labels = []
deaths = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[4]>0: 
        labels.append(row[0])   
        deaths.append(row[4])   

# Create pie chart
plt.figure(figsize=(10, 10))
plt.pie(deaths, labels=labels, autopct='%1.1f%%', startangle=140)
plt.title("Share of Total COVID-19 Deaths in South America")
plt.tight_layout()
plt.savefig("deaths_pie_chart_all.png")
plt.close()

# Create a new sheet for the chart
ws_pie = wb.create_sheet("Pie Chart")
img_pie = Image("deaths_pie_chart_all.png")
ws_pie.add_image(img_pie, "B2")


#----------------------------------------------------------------------------------------------
#Scatter Plot
tests = []
cases = []
countries = []

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[7] and row[3]:  
        tests.append(row[7])       
        cases.append(row[3])       
        countries.append(row[0])   
        
        
# Create scatter plot
plt.figure(figsize=(10, 6))
plt.scatter(tests, cases, color='darkorange')

# Add country labels
for i in range(len(countries)):
    plt.annotate(countries[i], (tests[i], cases[i]), fontsize=8)

plt.title("COVID-19 Tests vs Total Cases in South America")
plt.xlabel("Tests Conducted")
plt.ylabel("Total Cases")
plt.tight_layout()
plt.savefig("scatter_tests_vs_cases.png")
plt.close()


# Create a new sheet for the chart
ws_scatter = wb.create_sheet("Scatter Chart")
img_scatter = Image("scatter_tests_vs_cases.png")
ws_scatter.add_image(img_scatter, "B2")


#--------------------------------------------------------------------------------------------
# Save the workbook
wb.save("Module_8_Assignment.xlsx")
